from io import BytesIO

from openpyxl import load_workbook

from traceability_excel import (
    build_traceability_workbook,
    normalise_record,
    parse_traceability_workbook,
)


def test_bulk_workbook_round_trip_preserves_existing_record():
    content = build_traceability_workbook(
        {
            "raw": [{
                "id": "raw-1",
                "intakeDate": "2026-08-23",
                "materialName": "Sugar",
                "sweetdreamsBatchCode": "SD-100",
                "totalWeightKg": 25.5,
            }],
            "finished": [],
            "usage": [],
        },
        {"itemTypes": ["Ingredient"], "packagingTypes": ["Bag"]},
        ["raw", "finished", "usage"],
    )

    workbook = load_workbook(BytesIO(content), read_only=False, data_only=False)
    assert workbook.sheetnames[:4] == [
        "Instructions", "Lists", "Raw Material Intake", "Finished Batches"
    ]
    assert workbook["Lists"].sheet_state == "hidden"
    assert workbook["Raw Material Intake"].freeze_panes == "A2"
    workbook.close()

    parsed = parse_traceability_workbook(content)
    assert parsed["raw"][0]["id"] == "raw-1"
    assert parsed["raw"][0]["intakeDate"] == "2026-08-23"
    assert parsed["raw"][0]["totalWeightKg"] == 25.5


def test_normalise_record_requires_key_fields_and_typed_numbers():
    try:
        normalise_record("finished", {"productionDate": "2026-08-23"})
        assert False, "Expected validation error"
    except ValueError as error:
        assert "Finished Product" in str(error)

    record = normalise_record("usage", {
        "usageDate": "2026-08-23",
        "sweetdreamsBatchCode": "RAW-1",
        "finishedBatchCode": "FIN-1",
        "quantityUsedKg": "1,250.5",
    })
    assert record["quantityUsedKg"] == 1250.5


def test_uk_date_text_is_accepted_and_normalised_for_storage():
    record = normalise_record("raw", {
        "intakeDate": "01/09/2026", "materialName": "Sugar",
        "sweetdreamsBatchCode": "SD-100", "bestBeforeDate": "30/09/2026",
    })
    assert record["intakeDate"] == "2026-09-01"
    assert record["bestBeforeDate"] == "2026-09-30"
