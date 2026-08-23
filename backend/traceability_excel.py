"""Excel import/export helpers for company traceability records."""

from __future__ import annotations

from datetime import date, datetime
from io import BytesIO
from typing import Any, Iterable
from zipfile import ZipFile

from openpyxl import Workbook, load_workbook
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo


MAX_ROWS_PER_SHEET = 10_000
DEFAULT_CONFIG = {
    "itemTypes": ["Ingredient", "Packaging", "Additive"],
    "packagingTypes": ["Bag", "Box", "Pallet"],
}

TRACEABILITY_SCHEMAS = {
    "raw": {
        "sheet": "Raw Material Intake",
        "collection": "traceability_raw_intakes",
        "response_key": "rawIntakes",
        "date_field": "intakeDate",
        "required": {"intakeDate", "materialName", "sweetdreamsBatchCode"},
        "columns": [
            ("Record ID", "id", "text"),
            ("Intake Date", "intakeDate", "date"),
            ("Supplier Name", "supplierName", "text"),
            ("Material Name", "materialName", "text"),
            ("Best Before Date", "bestBeforeDate", "date"),
            ("Sweetdreams Batch Code", "sweetdreamsBatchCode", "text"),
            ("Supplier Batch Code", "supplierBatchCode", "text"),
            ("Pallet Number", "palletNumber", "text"),
            ("Number of Cases", "numberOfCases", "number"),
            ("Total Weight KG", "totalWeightKg", "number"),
            ("Item Type", "itemType", "text"),
            ("Packaging Type", "packagingType", "text"),
            ("Packaging SKU", "packagingSku", "text"),
            ("Units per Pallet", "unitsPerPallet", "number"),
        ],
    },
    "finished": {
        "sheet": "Finished Batches",
        "collection": "traceability_finished_batches",
        "response_key": "finishedBatches",
        "date_field": "productionDate",
        "required": {"productionDate", "finishedProduct", "finishedBatchCode"},
        "columns": [
            ("Record ID", "id", "text"),
            ("Production Date", "productionDate", "date"),
            ("Finished Product", "finishedProduct", "text"),
            ("Finished Batch Code", "finishedBatchCode", "text"),
            ("Units Produced", "unitsProduced", "number"),
            ("Line Number", "lineNumber", "text"),
            ("Best Before Date", "bestBeforeDate", "date"),
        ],
    },
    "usage": {
        "sheet": "Material Usage",
        "collection": "traceability_material_usage",
        "response_key": "materialUsage",
        "date_field": "usageDate",
        "required": {"usageDate", "sweetdreamsBatchCode", "finishedBatchCode"},
        "columns": [
            ("Record ID", "id", "text"),
            ("Usage Date", "usageDate", "date"),
            ("Sweetdreams Batch Code", "sweetdreamsBatchCode", "text"),
            ("Pallet Number", "palletNumber", "text"),
            ("Finished Batch Code", "finishedBatchCode", "text"),
            ("Quantity Used KG", "quantityUsedKg", "number"),
            ("Quantity Wasted KG", "quantityWastedKg", "number"),
            ("Units Used", "unitsUsed", "number"),
            ("Units Wasted", "unitsWasted", "number"),
        ],
    },
}


def _safe_text(value: Any) -> str:
    text = "" if value is None else str(value).strip()
    if text.startswith(("=", "+", "-", "@")):
        return "'" + text
    return text


def _date_value(value: Any) -> date | str | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    try:
        return date.fromisoformat(text[:10])
    except ValueError:
        return text


def _number_value(value: Any) -> float | int | None:
    if value in (None, ""):
        return None
    if isinstance(value, bool):
        raise ValueError("must be a number")
    if isinstance(value, (int, float)):
        return value
    number = float(str(value).replace(",", "").strip())
    return int(number) if number.is_integer() else number


def normalise_record(record_type: str, values: dict[str, Any]) -> dict[str, Any]:
    """Validate and normalize a traceability record from JSON or Excel."""
    if record_type not in TRACEABILITY_SCHEMAS:
        raise ValueError("Unknown traceability record type")
    schema = TRACEABILITY_SCHEMAS[record_type]
    result: dict[str, Any] = {}
    for _label, field, kind in schema["columns"]:
        value = values.get(field)
        if field == "id":
            if value not in (None, ""):
                result[field] = str(value).strip()
            continue
        if kind == "date":
            parsed = _date_value(value)
            if isinstance(parsed, date):
                result[field] = parsed.isoformat()
            elif parsed in (None, ""):
                result[field] = ""
            else:
                raise ValueError(f"{_label} must be a valid date")
        elif kind == "number":
            try:
                parsed_number = _number_value(value)
            except (TypeError, ValueError) as exc:
                raise ValueError(f"{_label} must be a number") from exc
            result[field] = "" if parsed_number is None else parsed_number
        else:
            result[field] = "" if value is None else str(value).strip()

    missing = [
        label
        for label, field, _kind in schema["columns"]
        if field in schema["required"] and result.get(field) in (None, "")
    ]
    if missing:
        raise ValueError("Missing required value(s): " + ", ".join(missing))
    return result


def build_traceability_workbook(
    records_by_type: dict[str, Iterable[dict[str, Any]]],
    config: dict[str, list[str]] | None,
    selected_types: Iterable[str],
) -> bytes:
    selected = [kind for kind in selected_types if kind in TRACEABILITY_SCHEMAS]
    if not selected:
        raise ValueError("Select at least one data type")

    config_values = DEFAULT_CONFIG if config is None else config
    merged_config = {
        "itemTypes": list(config_values.get("itemTypes", DEFAULT_CONFIG["itemTypes"])),
        "packagingTypes": list(config_values.get("packagingTypes", DEFAULT_CONFIG["packagingTypes"])),
    }
    workbook = Workbook()
    instructions = workbook.active
    instructions.title = "Instructions"
    instructions.sheet_view.showGridLines = False
    instructions["A1"] = "Traceability bulk workbook"
    instructions["A1"].font = Font(size=18, bold=True, color="1A7A6E")
    instructions["A3"] = "How to use this workbook"
    instructions["A3"].font = Font(size=12, bold=True)
    guidance = [
        "Existing records are included with a Record ID. Do not change that ID; rows with an existing ID are skipped during upload.",
        "To add records, enter them on the relevant sheet and leave Record ID blank.",
        "Required fields are highlighted in pale red. Dates should be entered as dates and numeric fields as numbers.",
        "Do not rename sheets or column headings. Upload the completed .xlsx file from Traceability > Bulk Upload.",
        "Valid rows are imported even if other rows contain errors; the upload result identifies each rejected row.",
    ]
    for row_number, line in enumerate(guidance, start=4):
        instructions.cell(row=row_number, column=1, value=line)
        instructions.cell(row=row_number, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    instructions.column_dimensions["A"].width = 115

    lists = workbook.create_sheet("Lists")
    lists["A1"] = "Item Types"
    lists["B1"] = "Packaging Types"
    for row_number, value in enumerate(merged_config["itemTypes"], start=2):
        lists.cell(row=row_number, column=1, value=_safe_text(value))
    for row_number, value in enumerate(merged_config["packagingTypes"], start=2):
        lists.cell(row=row_number, column=2, value=_safe_text(value))
    item_end = max(2, len(merged_config["itemTypes"]) + 1)
    packaging_end = max(2, len(merged_config["packagingTypes"]) + 1)
    named_ranges = (
        DefinedName("TraceabilityItemTypes", attr_text=f"'Lists'!$A$2:$A${item_end}"),
        DefinedName("TraceabilityPackagingTypes", attr_text=f"'Lists'!$B$2:$B${packaging_end}"),
    )
    for named_range in named_ranges:
        if hasattr(workbook.defined_names, "add"):
            workbook.defined_names.add(named_range)
        else:  # openpyxl 3.0 compatibility for local verification
            workbook.defined_names.append(named_range)

    header_fill = PatternFill("solid", fgColor="1A7A6E")
    required_fill = PatternFill("solid", fgColor="FCE8E6")
    for record_type in selected:
        schema = TRACEABILITY_SCHEMAS[record_type]
        worksheet = workbook.create_sheet(schema["sheet"])
        worksheet.sheet_view.showGridLines = False
        worksheet.freeze_panes = "A2"
        required_columns = []
        for column_number, (label, field, _kind) in enumerate(schema["columns"], start=1):
            cell = worksheet.cell(row=1, column=column_number, value=label)
            cell.fill = header_fill
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(vertical="center")
            if field in schema["required"]:
                required_columns.append(column_number)

        rows = list(records_by_type.get(record_type, []))
        for row_number, record in enumerate(rows, start=2):
            for column_number, (_label, field, kind) in enumerate(schema["columns"], start=1):
                value = record.get(field)
                if kind == "date":
                    value = _date_value(value)
                elif kind == "number":
                    try:
                        value = _number_value(value)
                    except (TypeError, ValueError):
                        value = None
                else:
                    value = _safe_text(value)
                cell = worksheet.cell(row=row_number, column=column_number, value=value)
                if kind == "date" and isinstance(value, date):
                    cell.number_format = "dd/mm/yyyy"
                elif kind == "number":
                    cell.number_format = "0.00"

        last_data_row = max(2, len(rows) + 1)
        for column_number in required_columns:
            for row_number in range(2, max(last_data_row + 25, 27)):
                worksheet.cell(row=row_number, column=column_number).fill = required_fill

        for column_number, (label, _field, kind) in enumerate(schema["columns"], start=1):
            width = min(max(len(label) + 3, 14), 30)
            if kind == "date":
                width = max(width, 15)
            worksheet.column_dimensions[worksheet.cell(1, column_number).column_letter].width = width

        table_ref = f"A1:{worksheet.cell(last_data_row, len(schema['columns'])).coordinate}"
        table = Table(displayName=f"Traceability{record_type.title()}", ref=table_ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False,
            showRowStripes=True, showColumnStripes=False,
        )
        worksheet.add_table(table)
        worksheet.auto_filter.ref = table_ref

        if record_type == "raw":
            item_col = next(i for i, (_l, f, _k) in enumerate(schema["columns"], 1) if f == "itemType")
            packaging_col = next(i for i, (_l, f, _k) in enumerate(schema["columns"], 1) if f == "packagingType")
            item_validation = DataValidation(type="list", formula1="TraceabilityItemTypes")
            packaging_validation = DataValidation(type="list", formula1="TraceabilityPackagingTypes")
            worksheet.add_data_validation(item_validation)
            worksheet.add_data_validation(packaging_validation)
            item_validation.add(
                f"{worksheet.cell(2, item_col).column_letter}2:"
                f"{worksheet.cell(MAX_ROWS_PER_SHEET + 1, item_col).column_letter}{MAX_ROWS_PER_SHEET + 1}"
            )
            packaging_validation.add(
                f"{worksheet.cell(2, packaging_col).column_letter}2:"
                f"{worksheet.cell(MAX_ROWS_PER_SHEET + 1, packaging_col).column_letter}{MAX_ROWS_PER_SHEET + 1}"
            )

    lists.sheet_state = "hidden"
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def parse_traceability_workbook(content: bytes) -> dict[str, list[dict[str, Any]]]:
    with ZipFile(BytesIO(content)) as archive:
        members = archive.infolist()
        if len(members) > 1_000 or sum(member.file_size for member in members) > 50 * 1024 * 1024:
            raise ValueError("Workbook expands beyond the safe processing limit")
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    parsed: dict[str, list[dict[str, Any]]] = {kind: [] for kind in TRACEABILITY_SCHEMAS}
    matched_sheets = 0
    for record_type, schema in TRACEABILITY_SCHEMAS.items():
        if schema["sheet"] not in workbook.sheetnames:
            continue
        matched_sheets += 1
        worksheet = workbook[schema["sheet"]]
        if worksheet.max_row > MAX_ROWS_PER_SHEET + 1:
            raise ValueError(f"{schema['sheet']} exceeds the {MAX_ROWS_PER_SHEET:,}-row limit")
        rows = worksheet.iter_rows(values_only=True)
        headers = next(rows, None)
        if not headers:
            continue
        header_indexes = {str(value).strip(): index for index, value in enumerate(headers) if value is not None}
        missing_headers = [label for label, _field, _kind in schema["columns"] if label not in header_indexes]
        if missing_headers:
            raise ValueError(f"{schema['sheet']} is missing column(s): {', '.join(missing_headers)}")
        for row_number, row in enumerate(rows, start=2):
            values = {
                field: row[header_indexes[label]] if header_indexes[label] < len(row) else None
                for label, field, _kind in schema["columns"]
            }
            if not any(value not in (None, "") for field, value in values.items() if field != "id"):
                continue
            try:
                record = normalise_record(record_type, values)
                record["__row__"] = row_number
                parsed[record_type].append(record)
            except ValueError as exc:
                parsed[record_type].append({"__row__": row_number, "__error__": str(exc)})
    workbook.close()
    if matched_sheets == 0:
        raise ValueError("Workbook does not contain any traceability data sheets")
    return parsed
